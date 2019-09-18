# -*- coding: utf-8 -*-
# !/usr/bin/env python3
#
# Copyright (c) 2017, Linaro Limited
# Copyright (c) 2018, Bobby Noelte
#
# SPDX-License-Identifier: Apache-2.0
#
# vim: ai:ts=4:sw=4
import os, fnmatch
import re
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
import time, datetime
import logging


def my_test_code_str():
    now = datetime.datetime.now()
    otherStyleTime = now.strftime("%Y_%m_%d_%H%M%S")
    print(otherStyleTime)
    xlsx_name = 'out\\' + otherStyleTime + 'test.xlsx'
    print(xlsx_name)
    wb = Workbook()  # 创建文件对象
    # grab the active worksheet
    ws = wb.active  # 获取第一个sheet
    # Data can be assigned directly to cells
    ws['A1'] = 42  # 写入数字
    ws['B1'] = "你好" + "automation test"  # 写入中文（unicode中文也可）
    # Rows can also be appended
    ws.append([1, 2, 3])  # 写入多个单元格
    ws['A2'] = datetime.datetime.now()  # 写入一个当前时间

    line = "36          8        456          0       1024        816   startup_stm32f746xx.o\n"
    b = line.split()
    ws.append(b)

    line = '0x080001d8   0x080001d8   0x00000000   Code   RO         2843    .ARM.Collect$$$$0000000F  mc_w.l(entry11a.o)\n'
    line = line[:-1]
    b = line.split()
    print(b)
    ws.append(b)
    wb.save(xlsx_name)


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False


def create_my_timename():
    now = datetime.datetime.now()
    otherStyleTime = now.strftime("%Y_%m_%d_%H%M%S")
    return otherStyleTime


def write_image_code(bbList, image_comm_name):
    xlsx_name = image_comm_name
    wb = Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet
    src_list = bbList

    for bb in src_list:
        ll = len(bb)
        if ll == 8:
            ws.column_dimensions['H'].width = 40
        if ll ==9:
            ws.column_dimensions['H'].width = 40
            ws.column_dimensions['I'].width = 40
        new_number = []
        for n in bb:
            if is_number(n) is True:
                new_number.append(int(n))
            else:
                new_number.append(n)
        bb = new_number
        ws.append(bb)
    # Save the file

    wb.save(xlsx_name)
    print('保存文件名 ' + xlsx_name)


def List_to_7Item(src_list):
    logging.debug(src_list)
    bb = len(src_list)
    list1 = src_list[:6]
    list2 = src_list[6:bb]
    s = ' '.join(list2)
    list1.append(s)
    logging.debug(list1)
    return list1


# Image component sizes
def open_image_component_sizes():
    # 第一种方法
    f = open("flash.map", "r")  # 设置文件对象
    src_line = f.readline()
    line = src_line[:-1]
    count = 1
    Image_start_flag = 0
    Image_com_list = []
    space_line = []
    item_cnt = 0
    while src_line:  # 直到读取完文件
        src_line = f.readline()  # 读取一行文件，包括换行符
        count += 1
        line = src_line[:-1]
        if len(line) < 3 or line.endswith('------') or line.startswith('====='):
            continue
        if 'Image component' in line:
            Image_start_flag = 1
            continue
        if Image_start_flag == 1:
            if 'Total R' in line:
                continue
            if 'Code (inc. data)' in line:
                Image_com_list.append(space_line)
                Image_com_list.append(space_line)
                result = ['Code(inc)', 'data', 'RO Data', 'RW Data', 'ZI Data', 'Debug', 'Object Name']
            else:
                # print('第%s行:%s' % (count, line))
                result = line.split()

                if line.endswith('Object Totals'):
                    Image_com_list.append(space_line)
                    Image_com_list.append(space_line)
                if len(result) > 7:
                    result = List_to_7Item(result)
            item_cnt += 1
            result.insert(0, item_cnt)
            Image_com_list.append(result)

    f.close()  # 关闭文件
    return Image_com_list


# Image memory map of image
def open_Memory_Map_of_image():
    # 第一种方法
    f = open("flash.map", "r")  # 设置文件对象
    src_line = f.readline()
    line = src_line[:-1]
    count = 1
    Image_start_flag = 0
    Image_list = []
    item_cnt = 0
    space_line = []
    while src_line:  # 直到读取完文件
        src_line = f.readline()  # 读取一行文件，包括换行符
        count += 1
        line = src_line[:-1]
        if len(line) < 3 or line.endswith('------') or line.startswith('====='):
            continue
        if 'Memory Map of the' in line:
            Image_start_flag = 1
            continue
        if 'Image component ' in line:
            break
        if Image_start_flag == 1:
            if 'Load Region LR_IROM1' in line:
                continue
            if 'Execution Region' in line:
                continue
            if 'Exec Addr' in line:
                result = []
                Image_list.append(space_line)
                Image_list.append(space_line)
                result = ['Exec Addr', 'Load Addr', 'Size', 'Type', 'Attr', 'Idx', 'E Section Name', 'Object']
            else:
                logging.debug('第%s行:%s' % (count, line))
                result = line.split()
                #logging.debug(result)
            item_cnt += 1
            result.insert(0, item_cnt)
            Image_list.append(result)

    f.close()  # 关闭文件
    logging.debug(Image_list)
    return Image_list


def create_image_comm_xlsx():
    image_comm_name = 'out\\' + create_my_timename() + '_image_comm.xlsx'
    image_com_list = open_image_component_sizes()
    write_image_code(image_com_list, image_comm_name)


def create_memory_map_xlsx():
    image_comm_name = 'out\\' + create_my_timename() + '_memory_map.xlsx'
    Image_list = open_Memory_Map_of_image()
    write_image_code(Image_list, image_comm_name)


def parse_arguments():
    rdh = argparse.RawDescriptionHelpFormatter
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=rdh)

    # parser.add_argument("-d", "--dts", required=True, help="DTS file")
    # parser.add_argument("-y", "--yaml", nargs='+', required=True,
    #                     help="YAML file directories, we allow multiple")
    # parser.add_argument("-i", "--include",
    #                     help="Generate include file for the build system")
    # parser.add_argument("-k", "--keyvalue",
    #                     help="Generate config file for the build system")
    # parser.add_argument("--old-alias-names", action='store_true',
    #                     help="Generate aliases also in the old way, without "
    #                          "compatibility information in their labels")
    return parser.parse_args()


def main():
    logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(funcName)s %(lineno)d %(message)s',
                        )
    logging.debug("test code debug")
    logging.info("test code info")
    args = parse_arguments()
    create_image_comm_xlsx()
    create_memory_map_xlsx()


if __name__ == '__main__':
    main()

'''
for n in src_list:
    tt += 1
    if tt > 7:
        logging.debug(" " + src_list[6])
        list_dst[6] = list_dst[6] + ' ' + n
    else:
        list_dst.append(n)
'''
