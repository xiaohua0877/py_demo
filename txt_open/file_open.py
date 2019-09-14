# -*- coding: utf-8 -*-
import re
from openpyxl import Workbook
from openpyxl import load_workbook


def test_code_str():
    wb = load_workbook('flash.xlsx')
    wb.guess_types = True
    ws = wb.active
    ws['A1'] = 42
    ws.append([1, 2, 3])
    line = "36          8        456          0       1024        816   startup_stm32f746xx.o\n"
    b = line.split()
    ws.append(b)
    print(b)
    wb.save("flash.xlsx")


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False


def test_code_write(b):
    wb = load_workbook('flash.xlsx')
    wb.guess_types = True
    ws = wb.active
    ws['A1'] = 42
    ws.append([1, 2, 3])
    src_list = b
    for bb in src_list:
        new_number = []
        for n in bb:
            if is_number(n) is True:
                new_number.append(int(n))
            else:
                new_number.append(n)
        bb = new_number
        ws.append(bb)
    wb.save("flash.xlsx")


def open_file_txt():
    # 第一种方法
    f = open("flash.map", "r")  # 设置文件对象
    src_line = f.readline()
    line = src_line[:-1]
    count = 1
    Image_start_flag = 0
    Image_com_list = []
    while src_line:  # 直到读取完文件
        src_line = f.readline()  # 读取一行文件，包括换行符
        count += 1
        line = src_line[:-1]
        if len(line) < 3 or line.endswith('------') or line.startswith('====='):
            continue
        if 'Image component' in line:
            Image_start_flag = 1
            continue
        if Image_start_flag == 1:
            if 'Total R' in line:
                continue
            if 'Code (inc. data)' in line:
                result = ['Code(inc)', 'data', 'RO Data', 'RW Data', 'ZI Data', 'Debug', 'Object Name']
            else:
                print('第%s行:%s' % (count, line))
                result = line.split()
                print(result)
            Image_com_list.append(result)

    f.close()  # 关闭文件
    return Image_com_list


# test_code_str()
b = open_file_txt()
test_code_write(b)

'''
#第二种方法
data = []
for line in open("data.txt","r"): #设置文件对象并读取每一行文件
    data.append(line)               #将每一行文件加入到list中


#第三种方法
f = open("data.txt","r")   #设置文件对象
data = f.readlines()  #直接将文件中按行读到list里，效果与方法2一样
f.close()             #关闭文件


print('第%s行:%s' % (count, line))
b = re.findall(r"\d+\.?\d*",line)
print(b)
pattern = re.compile(r'([a-z]*).o\n')
result = pattern.findall(line)
print(result)
'''
