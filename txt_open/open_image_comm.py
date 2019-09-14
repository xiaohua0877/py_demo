# -*- coding: utf-8 -*-
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import time, datetime


def my_test_code_str():
    now = datetime.datetime.now()
    otherStyleTime = now.strftime("%Y_%m_%d_ %H%M%S")
    print(otherStyleTime)
    xlsx_name = 'out\\' + otherStyleTime + 'test.xlsx'
    print(xlsx_name)
    wb = Workbook()  # 创建文件对象
    # grab the active worksheet
    ws = wb.active  # 获取第一个sheet
    # Data can be assigned directly to cells
    ws['A1'] = 42  # 写入数字
    ws['B1'] = "你好" + "automation test"  # 写入中文（unicode中文也可）
    # Rows can also be appended
    ws.append([1, 2, 3])  # 写入多个单元格
    ws['A2'] = datetime.datetime.now()  # 写入一个当前时间

    line = "36          8        456          0       1024        816   startup_stm32f746xx.o\n"
    b = line.split()
    ws.append(b)

    line = '0x080001d8   0x080001d8   0x00000000   Code   RO         2843    .ARM.Collect$$$$0000000F  mc_w.l(entry11a.o)\n'
    line = line[:-1]
    b = line.split()
    print(b)
    ws.append(b)

    # Save the file
    wb.save(xlsx_name)

    # wb = load_workbook('flash.xlsx')
    # wb.guess_types = True
    # ws = wb.active
    # ws['A1'] = 42
    # ws.append([1, 2, 3])

    # print(b)
    # wb.save("flash.xlsx")


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False


def create_my_timename():
    now = datetime.datetime.now()
    otherStyleTime = now.strftime("%Y_%m_%d_%H%M%S")
    return otherStyleTime


def write_image_code(bbList, image_comm_name):
    xlsx_name = image_comm_name
    wb = Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet
    src_list = bbList
    for bb in src_list:
        new_number = []
        for n in bb:
            if is_number(n) is True:
                new_number.append(int(n))
            else:
                new_number.append(n)
        bb = new_number
        ws.append(bb)
    # Save the file
    wb.save(xlsx_name)
    print('保存文件名 ' + xlsx_name)


def List_to_7Item(src_list):
    # print(src_list)
    tt = 0
    list_dst = []
    for n in src_list:
        tt += 1
        if tt > 7:
            # print(result1[6])
            list_dst[6] = list_dst[6] + ' ' + n
        else:
            list_dst.append(n)
    #print(list_dst)
    return list_dst


# Image component sizes
def open_image_component_sizes():
    # 第一种方法
    f = open("flash.map", "r")  # 设置文件对象
    src_line = f.readline()
    line = src_line[:-1]
    count = 1
    Image_start_flag = 0
    Image_com_list = []
    space_line = []
    while src_line:  # 直到读取完文件
        src_line = f.readline()  # 读取一行文件，包括换行符
        count += 1
        line = src_line[:-1]
        if len(line) < 3 or line.endswith('------') or line.startswith('====='):
            continue
        if 'Image component' in line:
            Image_start_flag = 1
            continue
        if Image_start_flag == 1:
            if 'Total R' in line:
                continue
            if 'Code (inc. data)' in line:
                Image_com_list.append(space_line)
                Image_com_list.append(space_line)
                result = ['Code(inc)', 'data', 'RO Data', 'RW Data', 'ZI Data', 'Debug', 'Object Name']
            else:
                #print('第%s行:%s' % (count, line))
                result = line.split()

                if line.endswith('Object Totals'):
                    Image_com_list.append(space_line)
                    Image_com_list.append(space_line)
                if len(result) > 7:
                    result = List_to_7Item(result)
            Image_com_list.append(result)

    f.close()  # 关闭文件
    return Image_com_list


# xlsx_name = 'out\\' + otherStyleTime + 'image_comm.xlsx'
image_comm_name = 'out\\' + create_my_timename() + '_image_comm.xlsx'
# my_test_code_str()
image_com_list = open_image_component_sizes()
write_image_code(image_com_list, image_comm_name)
