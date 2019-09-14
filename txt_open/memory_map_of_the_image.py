# -*- coding: utf-8 -*-
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import time, datetime


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False


def write_Memory_Map_data(bbList):
    now = datetime.datetime.now()
    otherStyleTime = now.strftime("%Y_%m_%d_%H%M%S")
    print(otherStyleTime)
    xlsx_name = 'out\\' + otherStyleTime + 'Memory_Map.xlsx'
    wb = Workbook()  # 创建文件对象
    ws = wb.active  # 获取第一个sheet
    ws['A1'] = 42
    ws.append([1, 2, 3])
    src_list = bbList
    for bb in src_list:
        new_number = []
        for n in bb:
            if is_number(n) is True:
                new_number.append(int(n))
            else:
                new_number.append(n)
        bb = new_number
        ws.append(bb)

    # Save the file
    wb.save(xlsx_name)


# Image memory map of image
def open_Memory_Map_of_image():
    # 第一种方法
    f = open("flash.map", "r")  # 设置文件对象
    src_line = f.readline()
    line = src_line[:-1]
    count = 1
    Image_start_flag = 0
    Image_list = []
    while src_line:  # 直到读取完文件
        src_line = f.readline()  # 读取一行文件，包括换行符
        count += 1
        line = src_line[:-1]
        if len(line) < 3 or line.endswith('------') or line.startswith('====='):
            continue
        if 'Memory Map of the' in line:
            Image_start_flag = 1
            continue
        if 'Image component ' in line:
            break
        if Image_start_flag == 1:
            if 'Load Region LR_IROM1' in line:
                continue
            if 'Execution Region' in line:
                continue
            if 'Exec Addr' in line:
                result = []
                Image_list.append(result)
                Image_list.append(result)
                result = ['Exec Addr', 'Load Addr', 'Size', 'Type', 'Attr', 'Idx', 'E Section Name', 'Object']
            else:
                print('第%s行:%s' % (count, line))
                result = line.split()
                #print(result)
            Image_list.append(result)

    f.close()  # 关闭文件
    print(Image_list)
    return Image_list


Image_list = open_Memory_Map_of_image()
write_Memory_Map_data(Image_list)

# Image_list = open_image_component_sizes()
# test_code_write(Image_list)
